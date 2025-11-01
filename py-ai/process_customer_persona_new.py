#!/usr/bin/env python3
import argparse
import csv
import json
import logging
import os
import random
import shutil
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openai import OpenAI  # type: ignore
except Exception:  # pragma: no cover
    OpenAI = None  # type: ignore

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

try:
    from openpyxl import load_workbook, Workbook  # type: ignore
except Exception as exc:  # pragma: no cover
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    print(str(exc), file=sys.stderr)
    sys.exit(1)


DEFAULT_INPUT_XLSX = "Customer Persona New.csv"
DEFAULT_OUTPUT_XLSX = "Customer Persona New.processed.csv"
DEFAULT_SYSTEM_PROMPT = "system_prompt.txt"
DEFAULT_USER_PROMPT = "user_prompt.txt"
DEFAULT_ATTACHMENT_XLSX = "customer title.xlsx"
DEFAULT_PROGRESS_PATH = "persona_progress.json"
DEFAULT_MODEL = os.environ.get("OPENAI_MODEL", "gpt-5")
DEFAULT_MAX_ROWS = int(os.environ.get("PERSONA_MAX_ROWS", "0"))


def detect_file_type(file_path: str) -> str:
    """Detect file type from extension."""
    ext = Path(file_path).suffix.lower()
    if ext in ('.csv',):
        return 'csv'
    elif ext in ('.xlsx', '.xls'):
        return 'xlsx'
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: .csv, .xlsx, .xls")


def read_csv_file(file_path: str) -> Tuple[List[str], List[List[str]]]:
    """Read CSV file and return (headers, rows)."""
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = list(reader)
    return headers, rows


def write_csv_file(file_path: str, headers: List[str], rows: List[List[str]]) -> None:
    """Write CSV file with headers and rows."""
    with open(file_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def setup_logging() -> None:
    level_name = os.environ.get("LOG_LEVEL", "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)
    logging.basicConfig(level=level, format="%(asctime)s %(levelname)s %(message)s")
    logging.getLogger("httpx").setLevel(logging.WARNING)


def read_text_file(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read().strip()


def load_attachment_as_csv_text(xlsx_path: str) -> str:
    if pd is None:
        raise RuntimeError("pandas is required to read attachment Excel. Install with: pip install pandas openpyxl")
    try:
        df = pd.read_excel(xlsx_path, engine="openpyxl")
    except Exception:
        df = pd.read_excel(xlsx_path)
    # Convert to CSV-style text to embed in prompt
    return df.to_csv(index=False)


def _extract_openai_content(resp: Any) -> str:
    """Extract text content from various OpenAI SDK response shapes.

    Supports:
    - ChatCompletions: resp.choices[0].message.content or resp["choices"][0]["message"]["content"]
    - Responses API: resp.output_text or first text block in resp.output
    - Dict responses or raw http responses with .json()
    """
    # 1) ChatCompletions object style
    try:
        choices = getattr(resp, "choices", None)
        if choices:
            # message.content (preferred)
            msg = getattr(choices[0], "message", None)
            if msg is not None:
                content = getattr(msg, "content", None)
                if isinstance(content, str) and content:
                    return content
            # text fallback
            txt = getattr(choices[0], "text", None)
            if isinstance(txt, str) and txt:
                return txt
    except Exception:
        pass

    # 2) New Responses API object
    try:
        output_text = getattr(resp, "output_text", None)
        if isinstance(output_text, str) and output_text:
            return output_text
        output = getattr(resp, "output", None)
        if output and isinstance(output, (list, tuple)):
            # Try to find first text segment
            first = output[0]
            content_list = getattr(first, "content", None) or first.get("content") if isinstance(first, dict) else None
            if content_list and isinstance(content_list, (list, tuple)):
                for part in content_list:
                    # SDK object may have .text; dict may have {"type":"output_text","text":...}
                    txt = getattr(part, "text", None) or (part.get("text") if isinstance(part, dict) else None)
                    if isinstance(txt, str) and txt:
                        return txt
    except Exception:
        pass

    # 3) Dict-like response
    try:
        if isinstance(resp, dict):
            if "choices" in resp and resp["choices"]:
                c0 = resp["choices"][0]
                if isinstance(c0, dict):
                    if "message" in c0 and isinstance(c0["message"], dict) and "content" in c0["message"]:
                        return str(c0["message"]["content"]) or ""
                    if "text" in c0:
                        return str(c0["text"]) or ""
            if "output_text" in resp and resp["output_text"]:
                return str(resp["output_text"]) or ""
    except Exception:
        pass

    # 4) Pydantic model fallback (Responses API object)
    try:
        model_dump = getattr(resp, "model_dump", None)
        if callable(model_dump):
            json_body = model_dump()
            if isinstance(json_body, dict):
                if "output_text" in json_body and json_body["output_text"]:
                    return str(json_body["output_text"]) or ""
                if "choices" in json_body and json_body["choices"]:
                    c0 = json_body["choices"][0]
                    if isinstance(c0, dict):
                        if "message" in c0 and isinstance(c0["message"], dict) and "content" in c0["message"]:
                            return str(c0["message"]["content"]) or ""
                        if "text" in c0:
                            return str(c0["text"]) or ""
    except Exception:
        pass

    # Fallback to string
    try:
        return str(resp)
    except Exception:
        return ""


def call_openai_with_retry(client: Any, model: str, messages: Any, max_retries: int = 5, timeout_seconds: int = 90) -> str:
    attempt = 0
    while True:
        try:
            resp = client.responses.create(
                model=model,
                input=messages,
                # max_tokens=1024,
                # temperature=0.2,
                # top_p=1.0,
                # timeout=timeout_seconds,
                # reasoning={"effort": "medium"},
            )
            # return _extract_openai_content(resp)
            return resp.output_text
        except Exception as e:
            attempt += 1
            if attempt > max_retries:
                logging.error("OpenAI call failed after %d attempts: %s", attempt - 1, e)
                raise
            sleep_seconds = min(2 ** attempt + random.uniform(0, 1), 30)
            logging.warning("OpenAI call failed (attempt %d). Retrying in %.2fs... %s", attempt, sleep_seconds, e)
            time.sleep(sleep_seconds)


def ensure_output_copied(input_xlsx: str, output_xlsx: str) -> None:
    if not os.path.exists(output_xlsx):
        shutil.copyfile(input_xlsx, output_xlsx)
        logging.info("Created output workbook by copying input: %s -> %s", input_xlsx, output_xlsx)


def map_headers(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        val = ("" if cell.value is None else str(cell.value)).strip()
        if val:
            header_map[val] = col_idx
    return header_map


def find_or_create_response_column(ws, header_name: str = "openai_response") -> int:
    header_map = map_headers(ws)
    if header_name in header_map:
        return header_map[header_name]
    # Find next empty column in header row
    col_idx = len(ws[1]) + 1
    ws.cell(row=1, column=col_idx, value=header_name)
    return col_idx


def load_progress(progress_path: str) -> Dict[str, Any]:
    if not os.path.exists(progress_path):
        return {}
    try:
        with open(progress_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_progress(progress_path: str, data: Dict[str, Any]) -> None:
    tmp_path = progress_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, progress_path)


def build_messages(system_prompt: str, user_prompt_template: str, new_persona: str, brand_name: str, attachment_csv_text: str) -> Any:
    user_prompt = (
        user_prompt_template
        .replace("{{New Persona}}", new_persona or "")
        .replace("{{Brand Name}}", brand_name or "")
    )
    # Attach the customer title file content to each prompt as CSV text
    attachment_block = (
        "ضمیمه فایل 'customer title.xlsx' (فرمت CSV):\n\n" + attachment_csv_text
    )
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
        {"role": "user", "content": attachment_block},
    ]
    return messages


def process_rows_csv(
    input_csv: str,
    output_csv: str,
    system_prompt_path: str,
    user_prompt_path: str,
    attachment_xlsx: str,
    start_row: int,
    progress_path: str,
    model: str,
    resume: bool,
    max_rows: Optional[int] = None,
) -> None:
    """Process CSV file with OpenAI."""
    if OpenAI is None:
        raise RuntimeError("The 'openai' package (v1.x) is required. Install with: pip install openai")
    if pd is None:
        raise RuntimeError("The 'pandas' package is required. Install with: pip install pandas openpyxl")

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Please set the OPENAI_API_KEY environment variable.")

    system_prompt = read_text_file(system_prompt_path)
    user_prompt_template = read_text_file(user_prompt_path)
    attachment_csv_text = load_attachment_as_csv_text(attachment_xlsx)

    # Read CSV
    headers, rows = read_csv_file(input_csv)

    # Find required columns
    if "New Persona" not in headers:
        raise RuntimeError("Input CSV must have a 'New Persona' column header.")
    brand_key = "brand_name" if "brand_name" in headers else ("Brand Name" if "Brand Name" in headers else None)
    if brand_key is None:
        raise RuntimeError("Input CSV must have a 'brand_name' (or 'Brand Name') column header.")

    persona_col_idx = headers.index("New Persona")
    brand_col_idx = headers.index(brand_key)

    # Add response column if not exists
    if "openai_response" not in headers:
        headers.append("openai_response")
    response_col_idx = headers.index("openai_response")
    output_rows = [list(row) + ([""] * (len(headers) - len(row))) for row in rows]

    # Progress tracking
    progress = load_progress(progress_path) if resume else {}
    last_processed_row = int(progress.get("last_processed_row", 0))
    if resume and last_processed_row >= 0:
        start_data_row = max(start_row - 1, last_processed_row + 1)
    else:
        start_data_row = max(start_row - 1, 0)

    client = OpenAI()
    processed = 0

    for row_idx in range(start_data_row, len(output_rows)):
        if max_rows is not None and max_rows > 0 and processed >= max_rows:
            logging.info("Reached max_rows=%d. Stopping.", max_rows)
            break

        row = output_rows[row_idx]
        new_persona = row[persona_col_idx] if persona_col_idx < len(row) else ""
        brand_name = row[brand_col_idx] if brand_col_idx < len(row) else ""

        if (not new_persona or not new_persona.strip()) and (not brand_name or not brand_name.strip()):
            logging.debug("Skipping empty row %d", row_idx + 1)
            last_processed_row = row_idx
            save_progress(progress_path, {"last_processed_row": last_processed_row, "output": output_csv})
            write_csv_file(output_csv, headers, output_rows)
            continue

        new_persona_str = str(new_persona).strip() if new_persona else ""
        brand_name_str = str(brand_name).strip() if brand_name else ""

        logging.info("Processing row %d/%d: brand='%s' persona_len=%d", row_idx + 1, len(output_rows), brand_name_str, len(new_persona_str))

        messages = build_messages(system_prompt, user_prompt_template, new_persona_str, brand_name_str, attachment_csv_text)

        try:
            answer = call_openai_with_retry(client, model, messages)
        except Exception as e:
            answer = f"ERROR: {e}"
            logging.exception("OpenAI call failed for row %d", row_idx + 1)

        # Expand row if necessary
        while len(row) <= response_col_idx:
            row.append("")
        row[response_col_idx] = answer

        write_csv_file(output_csv, headers, output_rows)
        last_processed_row = row_idx
        save_progress(progress_path, {"last_processed_row": last_processed_row, "output": output_csv})

        processed += 1

    logging.info("Done. Processed %d row(s). Results saved to %s. Progress in %s", processed, output_csv, progress_path)


def process_rows(
    input_xlsx: str,
    output_xlsx: str,
    system_prompt_path: str,
    user_prompt_path: str,
    attachment_xlsx: str,
    start_row: int,
    progress_path: str,
    model: str,
    resume: bool,
    max_rows: Optional[int] = None,
) -> None:
    """Process Excel file with OpenAI."""
    if OpenAI is None:
        raise RuntimeError("The 'openai' package (v1.x) is required. Install with: pip install openai")
    if pd is None:
        raise RuntimeError("The 'pandas' package is required. Install with: pip install pandas openpyxl")

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Please set the OPENAI_API_KEY environment variable.")

    system_prompt = read_text_file(system_prompt_path)
    user_prompt_template = read_text_file(user_prompt_path)
    attachment_csv_text = load_attachment_as_csv_text(attachment_xlsx)

    ensure_output_copied(input_xlsx, output_xlsx)

    wb = load_workbook(output_xlsx)
    ws = wb.active

    header_map = map_headers(ws)
    # Expecting columns: 'New Persona', 'brand_name'
    if "New Persona" not in header_map:
        raise RuntimeError("Input sheet must have a 'New Persona' column header.")
    # brand name might be 'brand_name' or 'Brand Name'
    brand_key = "brand_name" if "brand_name" in header_map else ("Brand Name" if "Brand Name" in header_map else None)
    if brand_key is None:
        raise RuntimeError("Input sheet must have a 'brand_name' (or 'Brand Name') column header.")

    response_col = find_or_create_response_column(ws, "openai_response")

    progress = load_progress(progress_path) if resume else {}
    last_processed_row = int(progress.get("last_processed_row", 0))  # sheet row index (1-based)
    if resume and last_processed_row >= 2:
        start_sheet_row = max(start_row, last_processed_row + 1)
    else:
        start_sheet_row = max(start_row, 2)  # ensure we start from data rows

    client = OpenAI()

    total_rows = ws.max_row
    processed = 0
    for sheet_row in range(start_sheet_row, total_rows + 1):
        if max_rows is not None and max_rows > 0 and processed >= max_rows:
            logging.info("Reached max_rows=%d. Stopping.", max_rows)
            break

        new_persona = ws.cell(row=sheet_row, column=header_map["New Persona"]).value
        brand_name = ws.cell(row=sheet_row, column=header_map[brand_key]).value

        # Skip empty rows
        if (new_persona is None or str(new_persona).strip() == "") and (brand_name is None or str(brand_name).strip() == ""):
            logging.debug("Skipping empty row %d", sheet_row)
            # Still update progress to avoid loops
            last_processed_row = sheet_row
            save_progress(progress_path, {"last_processed_row": last_processed_row, "response_column": response_col, "output": output_xlsx})
            continue

        new_persona_str = "" if new_persona is None else str(new_persona).strip()
        brand_name_str = "" if brand_name is None else str(brand_name).strip()

        logging.info("Processing row %d/%d: brand='%s' persona_len=%d", sheet_row, total_rows, brand_name_str, len(new_persona_str))

        messages = build_messages(system_prompt, user_prompt_template, new_persona_str, brand_name_str, attachment_csv_text)

        try:
            answer = call_openai_with_retry(client, model, messages)
        except Exception as e:
            answer = f"ERROR: {e}"
            logging.exception("OpenAI call failed for row %d", sheet_row)

        # Write response to the output workbook
        ws.cell(row=sheet_row, column=response_col, value=answer)

        # Save workbook and progress after each row to avoid losing results
        wb.save(output_xlsx)
        last_processed_row = sheet_row
        save_progress(progress_path, {"last_processed_row": last_processed_row, "response_column": response_col, "output": output_xlsx})

        processed += 1

    wb.save(output_xlsx)
    logging.info("Done. Processed %d row(s). Results saved to %s. Progress in %s", processed, output_xlsx, progress_path)


def parse_args(argv: Optional[list] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Process CSV or Excel file row-by-row with OpenAI, appending responses to output file while checkpointing progress.")
    parser.add_argument("--input-xlsx", default=DEFAULT_INPUT_XLSX, help=f"Input file (CSV or Excel, default: {DEFAULT_INPUT_XLSX})")
    parser.add_argument("--output-xlsx", default=DEFAULT_OUTPUT_XLSX, help=f"Output file (format matches input, default: {DEFAULT_OUTPUT_XLSX})")
    parser.add_argument("--system-prompt", default=DEFAULT_SYSTEM_PROMPT, help=f"System prompt file (default: {DEFAULT_SYSTEM_PROMPT})")
    parser.add_argument("--user-prompt", default=DEFAULT_USER_PROMPT, help=f"User prompt file with placeholders (default: {DEFAULT_USER_PROMPT})")
    parser.add_argument("--attachment-xlsx", default=DEFAULT_ATTACHMENT_XLSX, help=f"Attachment Excel to include in each prompt (default: {DEFAULT_ATTACHMENT_XLSX})")
    parser.add_argument("--progress", default=DEFAULT_PROGRESS_PATH, help=f"Progress checkpoint path (default: {DEFAULT_PROGRESS_PATH})")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"OpenAI model (default: {DEFAULT_MODEL})")
    parser.add_argument("--start-row", type=int, default=2, help="Starting Excel sheet row (1-based). Data typically starts at 2.")
    parser.add_argument("--no-resume", action="store_true", help="Do not resume from progress file; start at --start-row.")
    parser.add_argument("--max-rows", type=int, default=DEFAULT_MAX_ROWS, help=f"Max number of rows to process (<=0 means unlimited). Default: {DEFAULT_MAX_ROWS}")
    return parser.parse_args(argv)


def main(argv: Optional[list] = None) -> int:
    setup_logging()
    args = parse_args(argv)

    try:
        # Detect file type from input file
        input_path = os.path.abspath(args.input_xlsx)
        output_path = os.path.abspath(args.output_xlsx)
        file_type = detect_file_type(input_path)

        # Ensure output format matches input
        output_type = detect_file_type(output_path)
        if output_type != file_type:
            logging.warning("Output file type (%s) doesn't match input (%s). Using input format.", output_type, file_type)
            # Change output extension to match input
            base = os.path.splitext(output_path)[0]
            if file_type == 'csv':
                output_path = base + '.csv'
            else:
                output_path = base + '.xlsx'

        if file_type == 'csv':
            logging.info("Processing CSV file: %s", input_path)
            process_rows_csv(
                input_csv=input_path,
                output_csv=output_path,
                system_prompt_path=os.path.abspath(args.system_prompt),
                user_prompt_path=os.path.abspath(args.user_prompt),
                attachment_xlsx=os.path.abspath(args.attachment_xlsx),
                start_row=int(args.start_row),
                progress_path=os.path.abspath(args.progress),
                model=args.model,
                resume=not args.no_resume,
                max_rows=args.max_rows,
            )
        else:
            logging.info("Processing Excel file: %s", input_path)
            process_rows(
                input_xlsx=input_path,
                output_xlsx=output_path,
                system_prompt_path=os.path.abspath(args.system_prompt),
                user_prompt_path=os.path.abspath(args.user_prompt),
                attachment_xlsx=os.path.abspath(args.attachment_xlsx),
                start_row=int(args.start_row),
                progress_path=os.path.abspath(args.progress),
                model=args.model,
                resume=not args.no_resume,
                max_rows=args.max_rows,
            )
    except Exception as exc:
        logging.exception("Failed: %s", exc)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:])) 